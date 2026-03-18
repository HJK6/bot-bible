"""
Expense Manager — receipt storage, OCR, DynamoDB tracking, and tax report generation.

Usage:
    import sys
    sys.path.insert(0, "/Users/YOUR_USERNAME/telegram-claude-bot")
    from modules.expenses import ExpenseManager

    em = ExpenseManager()
    result = em.add_receipt("/path/to/receipt.jpg")  # OCR + store
    em.add_expense(amount=45.99, category="Office Expense", description="USB hub", date="2025-03-07")
    expenses = em.query(year="2025", category="Travel")
    em.export_csv("/tmp/expenses_2025.csv", year="2025")
    em.export_tax_excel("/tmp/Altum_2025_Expenses.xlsx", year="2025")
"""

import os
import re
import uuid
import csv
from datetime import datetime
from decimal import Decimal
from typing import Optional

import boto3
from boto3.dynamodb.conditions import Key
from PIL import Image
import pytesseract

EXPENSES_TABLE = "Expenses"
RECEIPTS_BUCKET = "YOUR_BOT_NAME-expense-receipts"

# Categories from the MyTaxFiler Business Organizer template
EXPENSE_CATEGORIES = [
    "Accounting",
    "Advertising",
    "Auto Expense",
    "Bank Charges",
    "Compensation of Officers",
    "Delivery",
    "Dues and Subscription",
    "Employee Benefit Programs",
    "Gifts",
    "Business Insurance",
    "Interest Paid",
    "Internet and Hosting",
    "Laundry",
    "Legal and Professional",
    "License",
    "Meals and Entertainment",
    "Miscellaneous",
    "Office Expense",
    "Outside Services/Contract Labor",
    "Parking and Tolls",
    "Pension and Profit-sharing",
    "Postage",
    "Printing",
    "Professional Development and Training",
    "Rents",
    "Repairs and Maintenance",
    "Salary and Wages",
    "Supplies",
    "Taxes",
    "Telephone",
    "Travel",
    "Utilities",
    "Home Office",
    "AWS Fees",
    "Business Mileage",
    "Other Expense",
]


class ExpenseManager:
    def __init__(self):
        dynamodb = boto3.resource("dynamodb", region_name="us-east-1")
        self.table = dynamodb.Table(EXPENSES_TABLE)
        self.s3 = boto3.client("s3", region_name="us-east-1")

    def ocr_receipt(self, image_path: str) -> dict:
        """Extract text from a receipt image using Tesseract OCR.
        Returns dict with 'text' (full OCR text) and 'amount' (parsed total or None).
        """
        img = Image.open(image_path)
        text = pytesseract.image_to_string(img)
        amount = self._parse_amount(text)
        return {"text": text, "amount": amount}

    def _parse_amount(self, text: str) -> Optional[float]:
        """Try to extract the total amount from OCR text.
        Looks for patterns like 'Total: $XX.XX', 'TOTAL $XX.XX', 'Amount: XX.XX', etc.
        """
        lines = text.strip().split("\n")
        # Try to find total/amount lines first (most reliable)
        total_patterns = [
            r"(?:total|amount\s*due|balance\s*due|grand\s*total|subtotal|amount)\s*[:=]?\s*\$?\s*([\d,]+\.?\d*)",
            r"\$\s*([\d,]+\.\d{2})",  # Any dollar amount as fallback
        ]

        # First pass: look for labeled totals
        for line in reversed(lines):  # totals are usually at the bottom
            line_lower = line.lower().strip()
            match = re.search(total_patterns[0], line_lower)
            if match:
                amount_str = match.group(1).replace(",", "")
                try:
                    return float(amount_str)
                except ValueError:
                    continue

        # Second pass: find the largest dollar amount (likely the total)
        all_amounts = []
        for line in lines:
            for match in re.finditer(total_patterns[1], line):
                try:
                    all_amounts.append(float(match.group(1).replace(",", "")))
                except ValueError:
                    continue

        if all_amounts:
            return max(all_amounts)

        return None

    def upload_receipt(self, file_path: str, expense_id: str) -> str:
        """Upload receipt image to S3. Returns the S3 key."""
        ext = os.path.splitext(file_path)[1] or ".jpg"
        s3_key = f"receipts/{expense_id}{ext}"
        self.s3.upload_file(file_path, RECEIPTS_BUCKET, s3_key)
        return s3_key

    def get_receipt_url(self, expense_id: str, s3_key: str) -> str:
        """Generate a presigned URL to view a receipt."""
        return self.s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": RECEIPTS_BUCKET, "Key": s3_key},
            ExpiresIn=3600,
        )

    def add_receipt(self, image_path: str, category: Optional[str] = None,
                    description: Optional[str] = None, date: Optional[str] = None) -> dict:
        """Process a receipt image: OCR, upload to S3, store in DynamoDB.
        Returns dict with expense_id, amount (or None if OCR failed), and ocr_text.
        The caller should confirm the amount with the user.
        """
        expense_id = str(uuid.uuid4())[:8]
        ocr_result = self.ocr_receipt(image_path)
        s3_key = self.upload_receipt(image_path, expense_id)

        expense_date = date or datetime.now().strftime("%Y-%m-%d")
        year = expense_date[:4]

        expense = {
            "expense_id": expense_id,
            "amount": Decimal(str(ocr_result["amount"])) if ocr_result["amount"] else None,
            "category": category or "Uncategorized",
            "description": description or "",
            "date": expense_date,
            "year": year,
            "receipt_s3_key": s3_key,
            "ocr_text": ocr_result["text"][:500],  # truncate for DynamoDB
            "status": "pending_confirmation" if ocr_result["amount"] else "needs_amount",
            "created_at": datetime.now().isoformat(),
        }
        # Remove None values for DynamoDB
        expense = {k: v for k, v in expense.items() if v is not None}
        self.table.put_item(Item=expense)

        return {
            "expense_id": expense_id,
            "amount": ocr_result["amount"],
            "ocr_text": ocr_result["text"],
            "s3_key": s3_key,
            "status": expense["status"],
        }

    def add_expense(self, amount: float, category: str, description: str,
                    date: Optional[str] = None, receipt_path: Optional[str] = None) -> dict:
        """Add an expense manually (no OCR). Optionally attach a receipt."""
        expense_id = str(uuid.uuid4())[:8]
        expense_date = date or datetime.now().strftime("%Y-%m-%d")
        year = expense_date[:4]

        expense = {
            "expense_id": expense_id,
            "amount": Decimal(str(amount)),
            "category": category,
            "description": description,
            "date": expense_date,
            "year": year,
            "status": "confirmed",
            "created_at": datetime.now().isoformat(),
        }

        if receipt_path:
            s3_key = self.upload_receipt(receipt_path, expense_id)
            expense["receipt_s3_key"] = s3_key

        self.table.put_item(Item=expense)
        return expense

    def confirm_amount(self, expense_id: str, amount: float) -> dict:
        """Confirm or correct the OCR-detected amount."""
        self.table.update_item(
            Key={"expense_id": expense_id},
            UpdateExpression="SET #a = :amount, #s = :status",
            ExpressionAttributeNames={"#a": "amount", "#s": "status"},
            ExpressionAttributeValues={":amount": Decimal(str(amount)), ":status": "confirmed"},
        )
        return {"expense_id": expense_id, "amount": amount, "status": "confirmed"}

    def update_expense(self, expense_id: str, **kwargs) -> dict:
        """Update fields on an expense (category, description, date, amount)."""
        update_parts = []
        values = {}
        names = {}
        for i, (key, val) in enumerate(kwargs.items()):
            alias = f"#k{i}"
            val_alias = f":v{i}"
            names[alias] = key
            values[val_alias] = Decimal(str(val)) if key == "amount" else val
            update_parts.append(f"{alias} = {val_alias}")

        if "year" not in kwargs and "date" in kwargs:
            update_parts.append("#yr = :yr")
            names["#yr"] = "year"
            values[":yr"] = kwargs["date"][:4]

        expr = "SET " + ", ".join(update_parts)
        self.table.update_item(
            Key={"expense_id": expense_id},
            UpdateExpression=expr,
            ExpressionAttributeNames=names,
            ExpressionAttributeValues=values,
        )
        return {"expense_id": expense_id, "updated": list(kwargs.keys())}

    def delete_expense(self, expense_id: str):
        """Delete an expense record."""
        expense = self.get_expense(expense_id)
        if expense.get("receipt_s3_key"):
            try:
                self.s3.delete_object(Bucket=RECEIPTS_BUCKET, Key=expense["receipt_s3_key"])
            except Exception:
                pass
        self.table.delete_item(Key={"expense_id": expense_id})

    def get_expense(self, expense_id: str) -> dict:
        """Get a single expense by ID."""
        resp = self.table.get_item(Key={"expense_id": expense_id})
        return resp.get("Item", {})

    def _query_index(self, index_name: str, key_name: str, key_value: str) -> list:
        """Query a GSI with automatic pagination."""
        items = []
        kwargs = {"IndexName": index_name, "KeyConditionExpression": Key(key_name).eq(key_value)}
        while True:
            resp = self.table.query(**kwargs)
            items.extend(resp.get("Items", []))
            if "LastEvaluatedKey" not in resp:
                break
            kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
        return items

    def _scan_all(self) -> list:
        """Scan entire table with pagination."""
        items = []
        kwargs = {}
        while True:
            resp = self.table.scan(**kwargs)
            items.extend(resp.get("Items", []))
            if "LastEvaluatedKey" not in resp:
                break
            kwargs["ExclusiveStartKey"] = resp["LastEvaluatedKey"]
        return items

    def query(self, year: Optional[str] = None, category: Optional[str] = None) -> list:
        """Query expenses by year and/or category."""
        if category:
            results = self._query_index("CategoryIndex", "category", category)
            if year:
                results = [e for e in results if e.get("year") == year]
            return sorted(results, key=lambda x: x.get("date", ""))
        elif year:
            return self._query_index("YearIndex", "year", year)
        else:
            return self._scan_all()

    def get_all(self, year: Optional[str] = None) -> list:
        """Get all expenses, optionally filtered by year."""
        if year:
            return self._query_index("YearIndex", "year", year)
        return self._scan_all()

    def summary(self, year: str) -> dict:
        """Get expense summary by category for a year."""
        expenses = self.query(year=year)
        by_category = {}
        total = 0
        for e in expenses:
            cat = e.get("category", "Uncategorized")
            amt = float(e.get("amount", 0))
            by_category.setdefault(cat, {"total": 0, "count": 0})
            by_category[cat]["total"] += amt
            by_category[cat]["count"] += 1
            total += amt
        return {"year": year, "total": total, "by_category": by_category, "expense_count": len(expenses)}

    def export_csv(self, output_path: str, year: Optional[str] = None) -> str:
        """Export expenses to CSV."""
        expenses = self.get_all(year=year)
        expenses.sort(key=lambda x: x.get("date", ""))

        with open(output_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Expense ID", "Date", "Category", "Description", "Amount", "Status"])
            for e in expenses:
                writer.writerow([
                    e.get("expense_id", ""),
                    e.get("date", ""),
                    e.get("category", ""),
                    e.get("description", ""),
                    e.get("amount", ""),
                    e.get("status", ""),
                ])
        return output_path

    def export_tax_excel(self, output_path: str, year: str,
                         business_name: str = "Altum Group",
                         income_1099: float = 0,
                         real_estate_commissions: float = 0) -> str:
        """Export expenses in the MyTaxFiler Business Organizer format.
        Generates the Profit & Loss section and per-category detail sheets.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        money_fmt = '#,##0.00'
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        expenses = self.get_all(year=year)
        expenses.sort(key=lambda x: x.get("date", ""))

        # Group by category
        by_category = {}
        for e in expenses:
            cat = e.get("category", "Uncategorized")
            by_category.setdefault(cat, []).append(e)

        # --- Business Profile (P&L) sheet ---
        ws = wb.active
        ws.title = "Business Profile"
        ws.column_dimensions["B"].width = 45
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 15

        ws["B1"] = "Business Information"
        ws["B1"].font = Font(bold=True, size=14)

        ws["B3"] = "Name of Business"
        ws["C3"] = business_name
        ws["C3"].fill = yellow_fill
        ws["B4"] = "Tax Year"
        ws["C4"] = int(year)
        ws["C4"].fill = yellow_fill

        ws["B6"] = f"{year} Profit and Loss"
        ws["B6"].font = Font(bold=True, size=12)
        ws["C7"] = "Amount"
        ws["C7"].font = bold
        ws["D7"] = "Notes"
        ws["D7"].font = bold

        # Income
        ws["B8"] = "Income"
        ws["B8"].font = bold
        ws["B9"] = "1099 Contracts"
        ws["C9"] = income_1099
        ws["C9"].number_format = money_fmt
        ws["C9"].fill = yellow_fill
        ws["B10"] = "Real Estate Commissions"
        ws["C10"] = real_estate_commissions
        ws["C10"].number_format = money_fmt
        ws["C10"].fill = yellow_fill
        ws["B11"] = "Total Income"
        ws["B11"].font = bold
        ws["C11"] = f"=SUM(C9:C10)"
        ws["C11"].number_format = money_fmt

        ws["B12"] = "Cost of Goods Sold"
        ws["C12"] = 0
        ws["C12"].number_format = money_fmt

        ws["B13"] = "Gross Income"
        ws["B13"].font = bold
        ws["C13"] = "=C11-C12"
        ws["C13"].number_format = money_fmt

        # Expenses
        ws["B15"] = "Expenses"
        ws["B15"].font = Font(bold=True, size=12)

        row = 16
        category_cells = {}
        for cat in EXPENSE_CATEGORIES:
            ws.cell(row=row, column=2, value=cat)
            cat_total = sum(float(e.get("amount", 0)) for e in by_category.get(cat, []))
            cell = ws.cell(row=row, column=3, value=cat_total if cat_total > 0 else None)
            cell.number_format = money_fmt
            if cat_total > 0:
                cell.fill = yellow_fill
            category_cells[cat] = row
            row += 1

        row += 1
        ws.cell(row=row, column=2, value="Total Expenses").font = bold
        ws.cell(row=row, column=3).value = f"=SUM(C16:C{row-2})"
        ws.cell(row=row, column=3).number_format = money_fmt
        total_exp_row = row

        row += 1
        ws.cell(row=row, column=2, value="Ordinary Income/Loss").font = bold
        ws.cell(row=row, column=3).value = f"=C13-C{total_exp_row}"
        ws.cell(row=row, column=3).number_format = money_fmt

        # Meals deduction note
        if "Meals and Entertainment" in by_category:
            row += 1
            ws.cell(row=row, column=2, value="Non-deductible (50% meals)")
            meals_row = category_cells.get("Meals and Entertainment", 0)
            if meals_row:
                ws.cell(row=row, column=3).value = f"=C{meals_row}/2"
                ws.cell(row=row, column=3).number_format = money_fmt

            row += 1
            ws.cell(row=row, column=2, value="Net Income/Loss").font = bold
            ws.cell(row=row, column=3).value = f"=C{total_exp_row+1}+C{row-1}"
            ws.cell(row=row, column=3).number_format = money_fmt

        # --- Detail sheets for categories with expenses ---
        for cat, cat_expenses in sorted(by_category.items()):
            if not cat_expenses:
                continue
            # Sheet name max 31 chars
            sheet_name = cat[:31]
            ws2 = wb.create_sheet(title=sheet_name)
            ws2.column_dimensions["A"].width = 50
            ws2.column_dimensions["B"].width = 15
            ws2.column_dimensions["C"].width = 15

            ws2["A1"] = "Description"
            ws2["B1"] = "Date"
            ws2["C1"] = "Amount"
            for col in ["A", "B", "C"]:
                ws2[f"{col}1"].font = header_font
                ws2[f"{col}1"].fill = header_fill
                ws2[f"{col}1"].border = thin_border

            for i, e in enumerate(cat_expenses, start=2):
                ws2.cell(row=i, column=1, value=e.get("description", ""))
                ws2.cell(row=i, column=2, value=e.get("date", ""))
                cell = ws2.cell(row=i, column=3, value=float(e.get("amount", 0)))
                cell.number_format = money_fmt

            total_row = len(cat_expenses) + 3
            ws2.cell(row=total_row, column=2, value="Total").font = bold
            ws2.cell(row=total_row, column=3).value = f"=SUM(C2:C{len(cat_expenses)+1})"
            ws2.cell(row=total_row, column=3).number_format = money_fmt
            ws2.cell(row=total_row, column=3).font = bold

        # --- All Expenses sheet ---
        ws_all = wb.create_sheet(title="All Expenses")
        ws_all.column_dimensions["A"].width = 12
        ws_all.column_dimensions["B"].width = 15
        ws_all.column_dimensions["C"].width = 30
        ws_all.column_dimensions["D"].width = 50
        ws_all.column_dimensions["E"].width = 15

        headers = ["Expense ID", "Date", "Category", "Description", "Amount"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_all.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for i, e in enumerate(expenses, start=2):
            ws_all.cell(row=i, column=1, value=e.get("expense_id", ""))
            ws_all.cell(row=i, column=2, value=e.get("date", ""))
            ws_all.cell(row=i, column=3, value=e.get("category", ""))
            ws_all.cell(row=i, column=4, value=e.get("description", ""))
            cell = ws_all.cell(row=i, column=5, value=float(e.get("amount", 0)))
            cell.number_format = money_fmt

        total_row = len(expenses) + 3
        ws_all.cell(row=total_row, column=4, value="Total").font = bold
        ws_all.cell(row=total_row, column=5).value = f"=SUM(E2:E{len(expenses)+1})"
        ws_all.cell(row=total_row, column=5).number_format = money_fmt
        ws_all.cell(row=total_row, column=5).font = bold

        wb.save(output_path)
        return output_path
